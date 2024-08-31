import datetime
import os
import re
from configparser import ConfigParser

import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(file, sheet_name):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    return sheet.max_column

def read_data_from_excel(file, sheet_name, row, column):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row, column).value

def write_data_into_excel(file, sheet_name, row, column, data):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    sheet.cell(row, column).value = data
    workbook.save(file)

def get_config(section, option):
    config = ConfigParser()
    file = os.path.join(os.path.dirname(os.path.abspath('.')), "test_data\\config.ini")
    config.read(file)
    return config.get(section, option)

def get_date(text):
    today = datetime.date.today()
    if str(text).find('hour') != -1:
        return today.strftime("%d %b %Y")
    elif str(text).find('day') != -1:
        digits = filter(str.isdigit, text)
        integer_str = int(''.join(digits))+ 1
        date_to_print = today - datetime.timedelta(days=integer_str)
        return date_to_print.strftime("%d %b %Y")
    elif str(text).find('week') != -1:
        digits = filter(str.isdigit, text)
        integer_str = (int(''.join(digits)) * 7) + 1
        date_to_print = today - datetime.timedelta(days=integer_str)
        return date_to_print.strftime("%d %b %Y")
    elif str(text).find('month') != -1:
        digits = filter(str.isdigit, text)
        integer_str = (int(''.join(digits)) * 30) + 1
        date_to_print = today - datetime.timedelta(days=integer_str)
        return date_to_print.strftime("%d %b %Y")
    elif str(text).find('year') != -1:
        digits = filter(str.isdigit, text)
        integer_str = (int(''.join(digits)) * 365) + 1
        date_to_print = today - datetime.timedelta(days=integer_str)
        return date_to_print.strftime("%d %b %Y")
    else:
        return text

def replace_special_character_with_space(value):
    return re.sub(r'[^a-zA-Z0-9]', ' ', value)

def split_string(input_string):
    # Use regular expression to split the string at the apostrophe
    parts = re.split(r"[’';]", input_string)
    # Strip leading and trailing spaces from each part
    parts = [part.strip() for part in parts]
    return parts

def split_string1(input_string):
    # Use regular expression to split the string at the apostrophe
    parts = re.split(r"[\\]", input_string)
    # Strip leading and trailing spaces from each part
    parts = [part.strip() for part in parts]
    return parts


file = os.path.join(os.path.dirname(os.path.abspath('.')), "test_data\\sample_file\\pdf-sample.pdf")
print(file)
print(split_string1(file))
